import io
import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG & STYLES
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Razorpay × Shopify", page_icon="🛡️", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #F8FAFC; }
    .main-header {
        background: #1F3864; padding: 25px; border-radius: 12px; 
        color: white; margin-bottom: 25px; text-align: center;
    }
    .metric-card {
        background: white; border-radius: 10px; padding: 15px;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: center;
        border-bottom: 4px solid #1F3864;
    }
    .metric-val { font-size: 24px; font-weight: 800; color: #1F3864; }
    .metric-lbl { font-size: 11px; color: #64748B; text-transform: uppercase; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE INITIALIZATION (Prevents disappearance on download)
# ─────────────────────────────────────────────────────────────────────────────
if 'rp_results' not in st.session_state:
    st.session_state.rp_results = None

# ─────────────────────────────────────────────────────────────────────────────
# FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def smart_read_excel(file_bytes, anchor_cols):
    try:
        preview = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=100)
        header_row_index = None
        anchors = [str(a).strip().lower() for a in anchor_cols]
        for i, row in preview.iterrows():
            row_values = [str(val).strip().lower() for val in row.values if pd.notna(val)]
            matches = sum(1 for a in anchors if any(a in val for val in row_values))
            if matches >= 2: 
                header_row_index = i
                break
        if header_row_index is None: return None
        df = pd.read_excel(io.BytesIO(file_bytes), header=header_row_index)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"Error reading file: {e}")
        return None

def get_excel_styles():
    thin = Side(style='thin', color='AAAAAA')
    return {
        'border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'al_c': Alignment(horizontal='center', vertical='center'),
        'hdr_fill': PatternFill('solid', fgColor='1F3864'),
        'hdr_font': Font(name='Arial', bold=True, color='FFFFFF', size=11),
        'credit_fill': PatternFill('solid', fgColor='E2EFDA'),
        'debit_fill': PatternFill('solid', fgColor='FCE4D6'),
        'tax_fill': PatternFill('solid', fgColor='FFF2CC')
    }

def create_single_excel(rows, headers, sheet_name, is_journal=False):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    styles = get_excel_styles()
    
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = styles['hdr_fill'], styles['hdr_font'], styles['al_c'], styles['border']
        ws.column_dimensions[get_column_letter(ci)].width = 25

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = styles['border']
            if is_journal:
                # Green for Credit (Sale), Orange for Debit (Refund)
                is_cr = "Receivable" not in str(row[1])
                cell.fill = styles['credit_fill'] if is_cr else styles['debit_fill']
            if ci == 1 and isinstance(val, (datetime.date, datetime.datetime)):
                cell.number_format = 'DD.MM.YYYY'
                
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# INTERFACE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-header"><h1>🛡️ Razorpay × Shopify Reconciliation</h1></div>', unsafe_allow_html=True)

u1, u2 = st.columns(2)
with u1: rp_file = st.file_uploader("Upload Razorpay Report", type=['xlsx'], key="rp")
with u2: sh_file = st.file_uploader("Upload Shopify Order Report", type=['xlsx'], key="sh_rp")

# Filename Inputs
st.subheader("📁 Custom Filenames")
fn1, fn2, fn3 = st.columns(3)
with fn1: j_name = st.text_input("Journal Name", f"Journal_RP_{datetime.date.today()}")
with fn2: t_name = st.text_input("Tax Journal Name", f"Tax_RP_{datetime.date.today()}")
with fn3: l_name = st.text_input("Lookup Name", f"Lookup_RP_{datetime.date.today()}")

if rp_file and sh_file:
    if st.button("▶ Run Process", type="primary", use_container_width=True):
        with st.spinner("Processing..."):
            rp_df_full = smart_read_excel(rp_file.getvalue(), ['transaction_entity', 'order_receipt', 'settled_at'])
            sh_df = smart_read_excel(sh_file.getvalue(), ['Order Number', 'Payment id'])

            if rp_df_full is not None and sh_df is not None:
                rp_df = rp_df_full[~rp_df_full['transaction_entity'].isin(['settlement.ondemand', 'adjustment'])].copy()
                sh_clean = sh_df.drop_duplicates('Payment id')
                email_map = sh_clean.set_index('Payment id')['Email'].to_dict() if 'Email' in sh_clean.columns else {}
                oid_map = sh_clean.set_index('Payment id')['Order Number'].to_dict()

                journal_rows = []
                lookup_rows = []
                
                for _, row in rp_df.iterrows():
                    receipt = str(row['order_receipt']).strip()
                    is_cr = float(row.get('credit', 0)) > 0
                    try: s_date = pd.to_datetime(row['settled_at']).date()
                    except: s_date = None
                    
                    email = email_map.get(receipt, "N/A")
                    order_no = oid_map.get(receipt, "N/A")
                    
                    # 1. Prepare Journal Data
                    journal_rows.append([
                        s_date, 
                        email if is_cr else 'Razorpay Payment Receivable', 
                        'Razorpay Payment Receivable' if is_cr else email, 
                        order_no, row['amount'], receipt
                    ])
                    
                    # 2. Prepare Lookup Data
                    lookup_rows.append([receipt, order_no, email, row['amount'], s_date])

                # Sort Journal: Credits first
                journal_rows.sort(key=lambda x: "Receivable" in str(x[1]))

                # 3. Prepare Tax Data
                total_tax_comm = round(rp_df_full['fee (exclusive tax)'].sum() + rp_df_full['tax'].sum(), 2)
                tax_rows = [[s_date, 'Razorpay Payment Receivable', 'Razorpay Commission Paid', '', total_tax_comm,]]

                # Store in Session State
                st.session_state.rp_results = {
                    'journal': create_single_excel(journal_rows, ['Order Date', 'Credit Account', 'Debit Account', 'Debit Reference No', 'gross Amount', 'Narration'], "Journal", True),
                    'tax': create_single_excel(tax_rows, ['Order Date', 'Credit Account', 'Debit Account', 'Debit Reference No', 'Amount', 'Narration'], "Tax Journal"),
                    'lookup': create_single_excel(lookup_rows, ['Razorpay ID', 'Order No', 'Email', 'Amount', 'Date'], "Lookup"),
                    'metrics': {
                        'total': len(journal_rows),
                        'tax': total_tax_comm,
                        'refunds': sum(1 for r in journal_rows if "Receivable" in str(r[1]))
                    }
                }

# ─────────────────────────────────────────────────────────────────────────────
# DOWNLOAD SECTION (Persists after clicking)
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.rp_results:
    res = st.session_state.rp_results
    st.divider()
    
    m1, m2, m3 = st.columns(3)
    m1.markdown(f'<div class="metric-card"><div class="metric-val">{res["metrics"]["total"]}</div><div class="metric-lbl">Total Txns</div></div>', unsafe_allow_html=True)
    m2.markdown(f'<div class="metric-card"><div class="metric-val">₹{res["metrics"]["tax"]}</div><div class="metric-lbl">Tax + Commission</div></div>', unsafe_allow_html=True)
    m3.markdown(f'<div class="metric-card"><div class="metric-val">{res["metrics"]["refunds"]}</div><div class="metric-lbl">Refunds</div></div>', unsafe_allow_html=True)
    
    st.write("")
    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button("⬇️ Download Journal", res['journal'], f"{j_name}.xlsx", use_container_width=True)
    with d2:
        st.download_button("⬇️ Download Tax Journal", res['tax'], f"{t_name}.xlsx", use_container_width=True)
    with d3:
        st.download_button("⬇️ Download Lookup", res['lookup'], f"{l_name}.xlsx", use_container_width=True)
