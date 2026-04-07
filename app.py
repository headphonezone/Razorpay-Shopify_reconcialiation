import streamlit as st
import pandas as pd
import re
import os
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.title("Excel Reconciliation Tool")

# ── Session state init ─────────────────────────────────────────────
if 'lookup_bytes'   not in st.session_state: st.session_state.lookup_bytes   = None
if 'journal_bytes'  not in st.session_state: st.session_state.journal_bytes  = None
if 'ready'          not in st.session_state: st.session_state.ready          = False

shopify_file  = st.file_uploader("Upload Shopify Excel",  type=["xlsx"])
razorpay_file = st.file_uploader("Upload Razorpay Excel", type=["xlsx"])

# ── Custom file names ──────────────────────────────────────────────
st.subheader("Output File Names")
lookup_filename  = st.text_input("Lookup File Name",  value="lookup",        placeholder="e.g. lookup_april")
journal_filename = st.text_input("Journal File Name", value="journal_final", placeholder="e.g. journal_april")

# ── Run button ─────────────────────────────────────────────────────
run_clicked = st.button("▶ Run Reconciliation", type="primary", disabled=not (shopify_file and razorpay_file))

if not shopify_file or not razorpay_file:
    st.info("Please upload both Shopify and Razorpay Excel files to continue.")
    st.session_state.ready = False

if run_clicked and shopify_file and razorpay_file:
    st.session_state.ready = False

    with st.spinner("Processing... please wait"):

        temp_dir      = tempfile.mkdtemp()
        shopify_path  = os.path.join(temp_dir, "shopify.xlsx")
        razorpay_path = os.path.join(temp_dir, "razorpay.xlsx")
        lookup_path   = os.path.join(temp_dir, "lookup_26.xlsx")
        output_path   = os.path.join(temp_dir, "journal_final_26.xlsx")

        with open(shopify_path, "wb") as f:
            f.write(shopify_file.getbuffer())
        with open(razorpay_path, "wb") as f:
            f.write(razorpay_file.getbuffer())

        shopify = pd.read_excel(shopify_path)
        rp      = pd.read_excel(razorpay_path)

        shopify['Payment id'] = shopify['Payment id'].astype(str).str.strip()
        rp['order_receipt']   = rp['order_receipt'].astype(str).str.strip()

        # ── Styles ────────────────────────────────────────────────────────────────
        def make_styles():
            header_font    = Font(name='Arial', bold=True, color='FFFFFF')
            header_fill    = PatternFill('solid', start_color='2F5496')
            center         = Alignment(horizontal='center', vertical='center')
            left           = Alignment(horizontal='left',   vertical='center')
            right          = Alignment(horizontal='right',  vertical='center')
            thin           = Side(style='thin', color='CCCCCC')
            border         = Border(left=thin, right=thin, top=thin, bottom=thin)
            alt_fill       = PatternFill('solid', start_color='EEF2FF')
            unmatched_fill = PatternFill('solid', start_color='FFE0E0')
            matched_fill   = PatternFill('solid', start_color='C6EFCE')
            mismatch_fill  = PatternFill('solid', start_color='FFEB9C')
            return header_font, header_fill, center, left, right, border, alt_fill, unmatched_fill, matched_fill, mismatch_fill

        def val_or_na(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return 'N/A'
            s = str(v).strip()
            return s if s else 'N/A'

        def write_headers(ws, headers, col_widths, header_font, header_fill, center, border):
            for col, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell           = ws.cell(row=1, column=col, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center
                cell.border    = border
                ws.column_dimensions[get_column_letter(col)].width = w
            ws.row_dimensions[1].height = 20

        def read_manual_edits():
            if not os.path.exists(lookup_path):
                return {}
            wb    = load_workbook(lookup_path, read_only=True, data_only=True)
            ws    = wb.active
            edits = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                pay_id = str(row[2]).strip() if row[2] else None
                if not pay_id or pay_id == 'N/A':
                    continue
                edits[pay_id] = {
                    'customer_name': row[0],
                    'order_no':      row[1],
                    'payment_id':    row[2],
                    'order_date':    row[3],
                    'amount':        row[4],
                    'fee':           row[5],
                    'tax':           row[6],
                    'gross_total':   row[7],
                    'dr_cr':         row[8],
                }
            wb.close()
            return edits

        def resolve(field, source_val, me):
            manual_val = me.get(field)
            if manual_val is not None and str(manual_val).strip() not in ('', 'N/A', 'None'):
                return manual_val
            return source_val

        # ── Shopify maps ──────────────────────────────────────────────────────────
        shopify_clean     = shopify.drop_duplicates('Payment id')
        shopify_email_map = shopify_clean.set_index('Payment id')['Email'].to_dict()
        shopify_oid_map   = shopify_clean.set_index('Payment id')['Order Number'].to_dict()
        shopify_pid_map   = {k: k for k in shopify_email_map}

        def extract_shopify_order_id(notes):
            if pd.isna(notes): return None
            match = re.search(r'"shopify_order_number":"([^"]+)"', str(notes))
            return match.group(1) if match else None

        rp['shopify_order_id']   = rp['payment_notes'].apply(extract_shopify_order_id)
        rp['customer_email']     = rp['order_receipt'].map(shopify_email_map)
        rp['shopify_order_no']   = rp['order_receipt'].map(shopify_oid_map)
        rp['shopify_payment_id'] = rp['order_receipt'].map(shopify_pid_map)

        manual_edits = read_manual_edits()

        header_font, header_fill, center, left, right, border, alt_fill, unmatched_fill, matched_fill, mismatch_fill = make_styles()

        # ── Lookup Sheet ──────────────────────────────────────────────────────────
        wb_lookup = Workbook()
        ws_lookup = wb_lookup.active
        ws_lookup.title = 'Lookup'

        lk_headers    = ['Customer Name', 'Order No', 'Payment ID', 'Order Date',
                         'Amount (Rs)', 'Fee (Rs)', 'Tax (Rs)', 'Gross Total (Rs)', 'DR/CR', 'Amount Check']
        lk_col_widths = [35, 20, 22, 14, 14, 10, 10, 16, 8, 14]

        write_headers(ws_lookup, lk_headers, lk_col_widths, header_font, header_fill, center, border)

        for i, row in rp.iterrows():
            r     = i + 2
            is_cr = row['credit'] > 0 and row['debit'] == 0
            dr_cr = 'CR' if is_cr else 'DR'

            receipt_key = val_or_na(row['order_receipt'])
            me          = manual_edits.get(receipt_key, {})

            email      = resolve('customer_name', val_or_na(row['customer_email']), me)
            order_no   = resolve('order_no',      val_or_na(row['shopify_order_no']), me)
            pay_id     = resolve('payment_id',    val_or_na(row['order_receipt']), me)
            amount     = resolve('amount',        row['amount'] if pd.notna(row['amount']) else 'N/A', me)
            fee        = resolve('fee',           row['fee (exclusive tax)'] if pd.notna(row['fee (exclusive tax)']) else 'N/A', me)
            tax        = resolve('tax',           row['tax'] if pd.notna(row['tax']) else 'N/A', me)
            gross      = resolve('gross_total',   row['credit'] if is_cr else row['debit'], me)
            dr_cr      = resolve('dr_cr',         dr_cr, me)

            try:
                raw_date   = resolve('order_date', row['entity_created_at'], me)
                order_date = pd.to_datetime(raw_date, dayfirst=True).date()
            except:
                order_date = val_or_na(row['entity_created_at'])

            try:
                calc      = round(float(amount) - float(fee) - float(tax), 2)
                actual    = round(float(gross), 2)
                amt_check = 'Matched' if calc == actual else f'Mismatch (expected {calc})'
            except:
                amt_check = 'N/A'

            if amt_check == 'Matched':
                row_fill = matched_fill
            elif 'Mismatch' in str(amt_check):
                row_fill = mismatch_fill
            else:
                row_fill = alt_fill if r % 2 == 0 else None

            values = [email, order_no, pay_id, order_date, amount, fee, tax, gross, dr_cr, amt_check]
            aligns = [left, center, left, center, right, right, right, right, center, center]

            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                cell           = ws_lookup.cell(row=r, column=col, value=val)
                cell.alignment = aln
                cell.border    = border
                if row_fill:
                    cell.fill = row_fill

        ws_lookup.freeze_panes = 'A2'
        wb_lookup.save(lookup_path)

        # ── Journal rows ────────────────────────────────
        journal_rows = []

        for i, row in rp.iterrows():
            is_cr = row['credit'] > 0 and row['debit'] == 0
            gross = row['amount'] if pd.notna(row['amount']) else 0   # ✅ CHANGED HERE

            receipt_key = val_or_na(row['order_receipt'])
            me          = manual_edits.get(receipt_key, {})

            narration_raw = val_or_na(row['shopify_payment_id'])
            if narration_raw == 'N/A':
                narration_raw = resolve('payment_id', val_or_na(row['order_receipt']), me)
            narration = narration_raw

            email    = resolve('customer_name', val_or_na(row['customer_email']), me)
            order_no = resolve('order_no',      val_or_na(row['shopify_order_no']), me)

            if is_cr:
                debit_acc  = 'Razorpay Payment Receivable'
                credit_acc = email
            else:
                debit_acc  = email
                credit_acc = 'Razorpay Payment Receivable'

            try:
                journal_date = pd.to_datetime(row['settled_at'], dayfirst=True).date()
            except:
                journal_date = None

            journal_rows.append({
                'is_cr':      is_cr,
                'order_date': journal_date,
                'credit_acc': credit_acc,
                'debit_acc':  debit_acc,
                'gross':      gross,
                'narration':  narration,
                'order_no':   order_no,
            })

        journal_rows.sort(key=lambda x: (0 if x['is_cr'] else 1))

        # ── Journal Sheet ─────────────────────────────────────────────────────────
        wb_journal = Workbook()
        ws_journal = wb_journal.active
        ws_journal.title = 'Journal'

        jn_headers    = ['Order Date', 'Credit Account', 'Debit Account', 'Debit Reference No',
                         'Gross Total (Rs)', 'Narration']
        jn_col_widths = [16, 35, 35, 18, 22, 32]

        write_headers(ws_journal, jn_headers, jn_col_widths, header_font, header_fill, center, border)

        cr_fill = PatternFill('solid', start_color='C6EFCE')
        dr_fill = PatternFill('solid', start_color='FFE0E0')

        for r_idx, entry in enumerate(journal_rows, start=2):
            row_fill = cr_fill if entry['is_cr'] else dr_fill
            values   = [
                entry['order_date'],
                entry['credit_acc'],
                entry['debit_acc'],
                entry['order_no'],
                entry['gross'],
                entry['narration'],
            ]
            aligns = [center, left, left, right, center, left]

            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                cell           = ws_journal.cell(row=r_idx, column=col, value=val)
                cell.border    = border
                cell.alignment = aln
                cell.fill      = row_fill

            ws_journal.cell(row=r_idx, column=1).number_format = 'DD/MM/YYYY'

            # Convert Order No to number
            ref_cell = ws_journal.cell(row=r_idx, column=4)
            try:
                ref_cell.value = int(float(str(entry['order_no']).strip()))
                ref_cell.number_format = '0'
            except:
                pass

        ws_journal.freeze_panes = 'A2'
        wb_journal.save(output_path)

    with open(lookup_path,  "rb") as f: st.session_state.lookup_bytes  = f.read()
    with open(output_path,  "rb") as f: st.session_state.journal_bytes = f.read()
    st.session_state.lookup_filename  = lookup_filename.strip()  or 'lookup'
    st.session_state.journal_filename = journal_filename.strip() or 'journal_final'
    st.session_state.ready = True

if st.session_state.ready:
    st.success("✅ Processing completed! Download your files below.")

    col1, col2 = st.columns(2)

    with col1:
        st.download_button(
            label="⬇ Download Lookup File",
            data=st.session_state.lookup_bytes,
            file_name=f"{st.session_state.lookup_filename}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_lookup"
        )

    with col2:
        st.download_button(
            label="⬇ Download Journal File",
            data=st.session_state.journal_bytes,
            file_name=f"{st.session_state.journal_filename}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_journal"
        )
